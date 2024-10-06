import sys
from pulp import COIN_CMD  # Import the COIN_CMD solver
import os  # Import os for file path operations
import pandas as pd
import numpy as np
import pulp
import math
from openpyxl import load_workbook
import folium
from folium import plugins
from branca.colormap import linear
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QTextEdit,
    QMessageBox, QProgressBar, QLabel
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtWebEngineWidgets import QWebEngineView  # Import QWebEngineView for displaying HTML content
from concurrent.futures import ThreadPoolExecutor

# Haversine formula to calculate the distance between two points
def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0  # Radius of the Earth in kilometers
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = R * c
    return distance

class WorkerThread(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal()

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        with ThreadPoolExecutor() as executor:
            future = executor.submit(self.process_excel)
            future.result()

    def process_excel(self):
        try:
            self.log_signal.emit("Starting processing...")
            # Read the Excel file
            self.log_signal.emit("Reading Excel file...")
            demand_df = pd.read_excel(self.file_path, sheet_name='Demand')
            stocks_df = pd.read_excel(self.file_path, sheet_name='Stocks')
            warehouse_df = pd.read_excel(self.file_path, sheet_name='Warehouse')

            # Add a dummy plant in warehouse_df
            dummy_location = pd.DataFrame([{'Branch': 'United Kingdom', 'Plant': 'UK01', 'Longitude': 0, 'Latitude': 0}])
            warehouse_df = pd.concat([dummy_location, warehouse_df], ignore_index=True)

            # Ensure 'Plant' and 'Storage Location' columns are strings
            demand_df['Plant'] = demand_df['Plant'].astype(str)
            demand_df['Storage Location'] = demand_df['Storage Location'].astype(str)
            stocks_df['Plant'] = stocks_df['Plant'].astype(str)
            stocks_df['Storage Location'] = stocks_df['Storage Location'].astype(str)

            # Create keys for Balance_to_serve and S_op
            demand_df['Key'] = demand_df['SKU'] + '&' + demand_df['Plant'] + '&' + demand_df['Storage Location']
            stocks_df['Key'] = stocks_df['SKU'] + '&' + stocks_df['Plant'] + '&' + stocks_df['Storage Location']

            # Initialize a list to store results
            results = []

            # Function to solve LP problem for given plants and resources
            def solve_lp(plants, initial_balance, S_op_array, Resource):
                lp_problem = pulp.LpProblem("Maximize_Min_Availability", pulp.LpMaximize)
                allocation_vars = pulp.LpVariable.dicts("Allocation", plants, lowBound=0, cat='Integer')
                min_availability = pulp.LpVariable("Min_Availability", lowBound=0)
                lp_problem += min_availability
                lp_problem += pulp.lpSum([allocation_vars[plant] for plant in plants]) <= Resource
                for i in range(len(plants)):
                    lp_problem += allocation_vars[plants[i]] <= initial_balance[i]
                    lp_problem += (1 - (initial_balance[i] - allocation_vars[plants[i]]) / S_op_array[i]) >= min_availability
                lp_problem.solve(COIN_CMD(path=os.path.join(sys._MEIPASS, 'cbc.exe') if hasattr(sys, '_MEIPASS') else 'cbc'))
                allocation = np.array([allocation_vars[plant].varValue for plant in plants])
                final_balance = initial_balance - allocation
                final_availability = 1 - (final_balance / S_op_array)
                return allocation, final_balance, final_availability

            # Iterate over each unique SKU
            total_skus = len(demand_df['SKU'].unique())
            for idx, sku in enumerate(demand_df['SKU'].unique(), 1):
                self.progress_signal.emit(int((idx / total_skus) * 50))  # Adjust progress scaling
                demand_sku_df = demand_df[demand_df['SKU'] == sku]
                stocks_sku_df = stocks_df[stocks_df['SKU'] == sku]
                Balance_to_serve = demand_sku_df.set_index('Key')['Balance_to_serve'].to_dict()
                S_op = demand_sku_df.set_index('Key')['S&Op'].to_dict()
                Resource = stocks_sku_df['Stocks'].sum()
                plants = list(Balance_to_serve.keys())
                initial_balance = np.array([Balance_to_serve[plant] for plant in plants])
                S_op_array = np.array([S_op[plant] for plant in plants])
                priority_array = demand_sku_df['Priority'].fillna(5).values
                allocation_total = np.zeros(len(plants))
                final_balance_total = initial_balance.copy()
                final_availability_total = (1 - (initial_balance / S_op_array)) * 100
                for priority in range(1, 6):
                    priority_plants = [plants[i] for i in range(len(plants)) if priority_array[i] == priority]
                    if priority_plants:
                        indices = [i for i in range(len(plants)) if priority_array[i] == priority]
                        allocation, final_balance, final_availability = solve_lp(
                            priority_plants,
                            final_balance_total[indices],
                            S_op_array[indices],
                            Resource
                        )
                        allocation_total[indices] += allocation
                        final_balance_total[indices] = final_balance
                        final_availability_total[indices] = final_availability * 100
                        Resource -= sum(allocation)
                result_df = demand_sku_df.copy()
                result_df['Allocation'] = allocation_total
                results.append(result_df)

            final_results_df = pd.concat(results)
            demand_df = final_results_df
            results = []
            distance_matrix_reference = []
            for idx, sku in enumerate(demand_df['SKU'].unique(), 1):
                self.progress_signal.emit(50 + int((idx / total_skus) * 30))  # Adjust progress scaling
                sku_demand_df = demand_df[demand_df['SKU'] == sku]
                sku_stocks_df = stocks_df[stocks_df['SKU'] == sku]
                dummy_stock = pd.DataFrame([{'Plant': 'UK01', 'Storage Location': '2000', 'Stocks': 1000000000000000, 'SKU': sku}])
                sku_stocks_df = pd.concat([dummy_stock, sku_stocks_df], ignore_index=True)
                sku_stocks_df['Key'] = sku_stocks_df['Plant'] + '_' + sku_stocks_df['Storage Location'].astype(str)
                sku_demand_df['Key'] = sku_demand_df['Plant'] + '_' + sku_demand_df['Storage Location'].astype(str)
                Stocks = {row['Key']: row['Stocks'] for _, row in sku_stocks_df.iterrows() if pd.notna(row['Key']) and pd.notna(row['Stocks'])}
                Demand = {row['Key']: row['Allocation'] for _, row in sku_demand_df.iterrows() if pd.notna(row['Key']) and pd.notna(row['Allocation'])}
                Locations = {}
                for _, row in warehouse_df.iterrows():
                    plant = row['Plant']
                    longitude = row['Longitude']
                    latitude = row['Latitude']
                    if pd.notna(plant) and pd.notna(longitude) and pd.notna(latitude):
                        Locations[plant] = (longitude, latitude)
                Distance_Matrix = {
                    (warehouse, plant): haversine(
                        Locations[warehouse.split('_')[0]][1],
                        Locations[warehouse.split('_')[0]][0],
                        Locations[plant.split('_')[0]][1],
                        Locations[plant.split('_')[0]][0]
                    )
                    for warehouse in Stocks
                    for plant in Demand
                }
                for (warehouse, plant), distance in Distance_Matrix.items():
                    from_plant = warehouse.split('_')[0]
                    to_plant = plant.split('_')[0]
                    from_branch = warehouse_df[warehouse_df['Plant'] == from_plant]['Branch'].values[0]
                    to_branch = warehouse_df[warehouse_df['Plant'] == to_plant]['Branch'].values[0]
                    distance_matrix_reference.append({
                        'From Plant': from_plant,
                        'From Branch': from_branch,
                        'To Plant': to_plant,
                        'To Branch': to_branch,
                        'Distance': distance
                    })
                prob = pulp.LpProblem("Minimize_Transportation_Distance", pulp.LpMinimize)
                routes = pulp.LpVariable.dicts("Route", Distance_Matrix, lowBound=0, cat='Continuous')
                prob += pulp.lpSum([Distance_Matrix[w_p] * routes[w_p] for w_p in Distance_Matrix]), "Total_Transportation_Distance"
                for warehouse in Stocks:
                    prob += pulp.lpSum([routes[(warehouse, plant)] for plant in Demand]) <= Stocks[warehouse], f"Stock_{warehouse}"
                for plant in Demand:
                    prob += pulp.lpSum([routes[(warehouse, plant)] for warehouse in Stocks]) == Demand[plant], f"Demand_{plant}"
                prob += pulp.lpSum([routes[w_p] for w_p in Distance_Matrix]) <= sum(Stocks.values()), "Total_Stock"
                prob.solve(COIN_CMD(path=os.path.join(sys._MEIPASS, 'cbc.exe') if hasattr(sys, '_MEIPASS') else 'cbc'))
                for w_p in Distance_Matrix:
                    allocation = routes[w_p].varValue
                    if allocation is not None and allocation > 0:
                        from_plant, from_storage = w_p[0].split('_')
                        to_plant, to_storage = w_p[1].split('_')
                        from_row = sku_stocks_df[sku_stocks_df['Key'] == w_p[0]].iloc[0]
                        to_row = sku_demand_df[sku_demand_df['Key'] == w_p[1]].iloc[0]
                        decision = "IBT"
                        if from_plant != to_plant:
                            if from_storage == "2000":
                                decision = "IBT"
                            else:
                                decision = "ICT and IBT"
                        else:
                            decision = "ICT"
                        results.append({
                            'SKU': sku,
                            'From Region': from_row['Region'],
                            'From Plant': from_plant,
                            'From Storage Location': from_storage,
                            'From Branch': warehouse_df[warehouse_df['Plant'] == from_plant]['Branch'].values[0],
                            'To Region': to_row['Region'],
                            'To Plant': to_plant,
                            'To Storage Location': to_storage,
                            'To Branch': warehouse_df[warehouse_df['Plant'] == to_plant]['Branch'].values[0],
                            'Quantity': allocation,
                            'Distance': Distance_Matrix[w_p],
                            'Decision': decision
                        })

            results_df = pd.DataFrame(results)
            results_df = results_df[results_df['From Plant'] != 'UK01']
            results_df['Key'] = results_df['SKU'] + '&' + results_df['To Plant'] + '&' + results_df['To Storage Location']
            results_df = results_df.merge(demand_df[['Key', 'Balance_to_serve', 'S&Op']], on='Key', how='left')
            results_df['Initial Availability %'] = 100 * (1 - (results_df['Balance_to_serve'] / results_df['S&Op']))
            results_df['Final Quantity'] = results_df.groupby('Key')['Quantity'].transform('sum')
            results_df['Final Availability %'] = 100 * (1 - ((results_df['Balance_to_serve'] - results_df['Final Quantity']) / results_df['S&Op']))
            results_df.drop(columns=['Final Quantity'], inplace=True)
            demand_df.drop(columns=['Key'], inplace=True)
            stocks_df.drop(columns=['Key'], inplace=True)
            results_df.drop(columns=['Key'], inplace=True)
            distance_matrix_df = pd.DataFrame(distance_matrix_reference).drop_duplicates()

            master_df = pd.read_excel(self.file_path, sheet_name='Master')
            if 'SKU Volume' not in master_df.columns or 'SKU Weight' not in master_df.columns:
                self.log_signal.emit("SKU Volume or SKU Weight columns are missing in the Master sheet.")
                return
            results_df = results_df.merge(master_df[['SKU', 'SKU Volume', 'SKU Weight']], on='SKU', how='left')
            results_df['Volume'] = results_df['Quantity'] * results_df['SKU Volume']
            results_df['Weight'] = results_df['Quantity'] * results_df['SKU Weight']
            results_df.drop(columns=['SKU Volume', 'SKU Weight'], inplace=True)
            
            # ----------------- Modified Dispatch Plan Generation Start -----------------
            
            # Vehicle capacity of each vehicle type
            vehicle_capacity = {
                '20_Feet': {'Volume': 1280, 'Weight': 6500},
                '32_Feet': {'Volume': 2040, 'Weight': 10000},
                'PTL': {'Volume': 1279, 'Weight': 6499}
            }

            # Function to select vehicles based on dispatches and vehicle capacities
            def select_vehicles(dispatches, vehicle_capacity):
                # Sort vehicle types by volume capacity in descending order
                sorted_vehicle_types = sorted(vehicle_capacity.items(), key=lambda x: x[1]['Volume'], reverse=True)
                
                results = []
                
                for dispatch in dispatches:
                    volume = dispatch['Volume']
                    weight = dispatch['Weight']
                    allocated_vehicles = {vehicle_type: 0 for vehicle_type in vehicle_capacity.keys()}

                    for vehicle_type, capacity in sorted_vehicle_types:
                        vehicle_volume = capacity['Volume']
                        vehicle_weight = capacity['Weight']
                        
                        num_vehicles = volume // vehicle_volume
                        if num_vehicles > 0:
                            allocated_vehicles[vehicle_type] += num_vehicles
                            volume -= num_vehicles * vehicle_volume
                            weight -= num_vehicles * vehicle_weight

                        if volume <= 0:
                            break

                    # Handle remaining volume with the smallest vehicle
                    if volume > 0:
                        smallest_vehicle_type, smallest_capacity = sorted_vehicle_types[-1]
                        if weight <= smallest_capacity['Weight']:
                            allocated_vehicles[smallest_vehicle_type] += 1
                            volume = 0

                    result = {
                        'From Plant': dispatch['From Plant'],
                        'To Plant': dispatch['To Plant'],
                        'Quantity': dispatch['Quantity'],
                        'Volume': dispatch['Volume'],
                        'Weight': dispatch['Weight']
                    }
                    result.update(allocated_vehicles)
                    results.append(result)
                
                return results

            # Create pivot table
            pivot_table = pd.pivot_table(
                results_df,
                values=['Quantity', 'Volume', 'Weight'],
                index=['From Plant', 'To Plant'],
                aggfunc={'Quantity': 'sum', 'Volume': 'sum', 'Weight': 'sum'}
            ).reset_index()

            # Convert pivot table to list of dispatches
            dispatches = pivot_table.to_dict('records')

            # Get the vehicle allocation results
            vehicle_allocations = select_vehicles(dispatches, vehicle_capacity)

            # Create a DataFrame from the results
            dispatch_plan_df = pd.DataFrame(vehicle_allocations)

            # ----------------- Modified Dispatch Plan Generation End -----------------

            # Initialize ExcelWriter
            with pd.ExcelWriter('Result.xlsx', engine='openpyxl') as writer:
                demand_df.to_excel(writer, sheet_name='Demand', index=False)
                stocks_df.to_excel(writer, sheet_name='Stocks', index=False)
                results_df.to_excel(writer, sheet_name='Solution', index=False)
                distance_matrix_df.to_excel(writer, sheet_name='Distance Matrix', index=False)
                dispatch_plan_df.to_excel(writer, sheet_name='Dispatch Plan', index=False)  # Updated line
                with pd.ExcelFile(self.file_path) as reader:
                    for sheet_name in ['Vehicle', 'Master', 'Warehouse']:
                        df = pd.read_excel(reader, sheet_name=sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = load_workbook('Result.xlsx')
            sheets = workbook.sheetnames
            sheets.insert(sheets.index('Solution') + 1, sheets.pop(sheets.index('Dispatch Plan')))
            workbook._sheets = [workbook[sheet] for sheet in sheets]
            workbook.save('Result.xlsx')
            self.log_signal.emit("Final Results and Dispatch Plan saved to Result.xlsx")
            
            # Generate the Folium map
            m = folium.Map(location=[20.5937, 78.9629], zoom_start=5)  # Centered on India
            unique_plants = pivot_table['From Plant'].unique()
            colormap = linear.Set1_09.scale(0, len(unique_plants)).to_step(len(unique_plants))
            color_map = {plant: colormap(i) for i, plant in enumerate(unique_plants)}
            for _, row in pivot_table.iterrows():
                from_plant = row['From Plant']
                to_plant = row['To Plant']
                from_location = warehouse_df[warehouse_df['Plant'] == from_plant][['Latitude', 'Longitude']].values[0]
                to_location = warehouse_df[warehouse_df['Plant'] == to_plant][['Latitude', 'Longitude']].values[0]
                line = folium.PolyLine(
                    locations=[from_location, to_location],
                    color=color_map[from_plant],
                    weight=2.5,
                    opacity=1
                ).add_to(m)
                plugins.PolyLineTextPath(
                    line,
                    '→→',
                    repeat=False,
                    offset=7,
                    attributes={'fill': color_map[from_plant], 'font-weight': 'bold', 'font-size': '35'}
                ).add_to(m)
            m.save('Network.html')  # This will overwrite any existing Network.html
            self.log_signal.emit("Network map saved to Network.html")
            self.progress_signal.emit(80)  # Update progress to 80%

            self.progress_signal.emit(100)  # Final progress update

            # Notify that the map has been saved
            self.log_signal.emit("Processing completed successfully.")
            
        except Exception as e:
            self.log_signal.emit(f"An error occurred: {str(e)}")
            self.progress_signal.emit(0)
        self.finished_signal.emit()

class MapWindow(QWidget):
    """A separate window to display the Network.html map."""
    def __init__(self, map_path):
        super().__init__()
        self.setWindowTitle("Network Map")
        self.setGeometry(150, 150, 1000, 800)  # Increased window size for better visibility
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        
        self.view = QWebEngineView()
        self.view.setUrl(QUrl.fromLocalFile(map_path))
        self.layout.addWidget(self.view)

class ExcelProcessorGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
    def set_default_message(self):
        self.log_text.setText("<b style='color: blue;'>Problem solving approach-</b><br>"
                                "<b style='color:  #004282;'>Objective Function-</b><br>1. Maximization of Minimum Availability across all plants basis priority,2.Minimization of Transportation distance,3. Minimization of Total Vehicle loadings<br>"
                                "<b style='color:  #004282;'>Constraints-</b><br>1.Allocation can not exceed available stocks,2.The total quantity allocated to each plant Can not exceed Balance_to_serve,3. if No Balance_to_serve cross RWH stock will be kept in CWH,4.Allocations will be served by nearest CWH<br>"
                                "<b style='color: red;'>Procedure to Use Tool:-</b><br>"
                                "1. Fill data in Input.elsx file,Enter Plant-SKU-Channel wise Balance_to_Serve and S&Op, Priority (1,2,3,4, 5- Default if Blank).<br>"
                                "2. Enter Stock across CWH,Maintain Longitude and Latitude in Warehouse sheet for all warehouses.<br>"
                                "3. Browse input.xlsx file and Click on 'Run Processing', Output will be saved in 'Result.xlsx'<br>"
                                "4. Network map will be saved in 'Network.html'<br>"
                                "5. Blank cells not allowed in input excel, especially Demand & Stock columns.Do not change any table headers in Input File.<br>"
                                "<br>"
                                "Please contact <b>Mahesh Pol(8149598176)</b> for any errors/bugs/logic modifications")

    def init_ui(self):
        self.setWindowTitle('Crompton Distribution Network Optimization⭐')
        self.setGeometry(100, 100, 800, 600)
        self.layout = QVBoxLayout()
        
        # Add heading
        self.heading_label = QLabel('<span style="font-size: 32px; font-weight: bold; color: #2E8B57;">Distribution Network Optimization</span>', self)
        self.heading_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.heading_label)
        
        self.browse_button = QPushButton('📂 Browse Excel File')
        self.browse_button.setStyleSheet("font-size: 20px; background-color: #2E8B57; color: white;")
        self.browse_button.clicked.connect(self.browse_file)
        self.layout.addWidget(self.browse_button)
        
        self.run_button = QPushButton('🧠 Run Processing')
        self.run_button.setStyleSheet("font-size: 20px; background-color: #004282; color: white;")
        self.run_button.clicked.connect(self.run_processing)
        self.layout.addWidget(self.run_button)
        
        self.view_map_button = QPushButton('🗺️ View Network Map')  # New button for viewing the map
        self.view_map_button.setStyleSheet("font-size: 20px; background-color: #FF8C00; color: white;")
        self.view_map_button.clicked.connect(self.view_map)
        self.view_map_button.setEnabled(False)  # Initially disabled
        self.layout.addWidget(self.view_map_button)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #004282;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #2E8B57;
                width: 20px;
            }
        """)
        self.layout.addWidget(self.progress_bar)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFixedSize(780, 350)
        self.layout.addWidget(self.log_text)
        
        # self.developed_by_label = QLabel('Developed by Crompton SCM Team<br>'
        # '<b style="color: blue;"><i>Transforming Distribution with Intelligent Planning🚀</i></b>')
        self.developed_by_label = QLabel('<span style="line-height: 2;">Developed by Crompton SCM Team</span><br>'
    '<b style="color: blue; line-height: 2;"><i>Intelligent Planning for Maximum Availability & Optimized Network at Perfect Time, Every Time.🚀</i></b>')

        self.developed_by_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.developed_by_label)
        
        self.setLayout(self.layout)
        
        # Set default message
        self.set_default_message()

    def browse_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Select Excel File", 
            "", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )
        if file_path:
            self.file_path = file_path
            self.log_text.append(f"Selected file: {file_path}")
            self.view_map_button.setEnabled(False)  # Disable map view until processing is done

    def run_processing(self):
        if hasattr(self, 'file_path'):
            self.view_map_button.setEnabled(False)  # Disable the map view button during processing
            self.worker_thread = WorkerThread(self.file_path)
            self.worker_thread.log_signal.connect(self.update_log)
            self.worker_thread.progress_signal.connect(self.update_progress)
            self.worker_thread.finished_signal.connect(self.processing_finished)
            self.worker_thread.start()
            self.log_text.clear()  # Clear the message box
        else:
            QMessageBox.warning(self, "No File Selected", "Please select an Excel file first.")

    def update_log(self, message):
        self.log_text.append(message)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def processing_finished(self):
        QMessageBox.information(self, "Processing Finished", "Excel processing has been completed.")
        # Check if Network.html exists before enabling the map view button
        map_path = os.path.abspath('Network.html')
        if os.path.exists(map_path):
            self.view_map_button.setEnabled(True)
        else:
            self.log_text.append("Warning: Network.html was not found. Please ensure it was generated correctly.")

    def view_map(self):
        try:
            map_path = os.path.abspath('Network.html')
            if not os.path.exists(map_path):
                QMessageBox.warning(
                    self, 
                    "Map Not Found", 
                    "Network.html not found. Please run the processing to generate the map."
                )
                return
            self.map_window = MapWindow(map_path)
            self.map_window.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load the map: {str(e)}")

def main():
    app = QApplication(sys.argv)
    gui = ExcelProcessorGUI()
    gui.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
